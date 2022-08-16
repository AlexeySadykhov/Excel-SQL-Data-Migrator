import sys
import os
import sqlite3
from sqlite3 import Error
import xlsxwriter as xlsx
import openpyxl as pyxl


def connect_to_db(path):
    try:
        connect = sqlite3.connect(path)
        return connect
    except Error as e:
        print(f"The error '{e}' occurred.")
        sys.exit(1)


def select_request(connect, request):
    cursor = connect.cursor()
    try:
        cursor.execute(request)
        result = cursor.fetchall()
        return result
    except Error as e:
        print(f"The error '{e}' occurred.")
        sys.exit(1)


def change_request(connect, request):
    cursor = connect.cursor()
    try:
        cursor.execute(request)
        connection.commit()
    except Error as e:
        print(f"The error '{e}' occurred.")
        sys.exit(1)


def create_query(tbl, cols):
    col_t = tuple(map(lambda x: " ".join(str(y) for y in [x, 'TEXT']), cols))
    col_s = str(col_t).replace("'", "")
    return f"CREATE TABLE {tbl} {str(col_s)};"


def insert_query(tbl, cols, vals):
    cls = ", ".join(str(x) for x in cols)
    vls = str(tuple(vals)).replace("None", "NULL")
    return f"INSERT INTO {tbl}({cls}) VALUES{vls};"


operation = input("""What type of conversion do you want to do 
(sqlite -> Excel / Excel -> sqlite):""")
if operation == 'sqlite -> Excel':
    in_file_location = input('Enter sqlite file location:')
    if os.path.isfile(in_file_location) is False:
        print('There is no such file.')
        sys.exit(1)
    else:
        connection = connect_to_db(in_file_location)
        table = input('Enter table name you want to export:')
        col_names = select_request(connection, f"SELECT name FROM PRAGMA_TABLE_INFO('{table}')")
        data = select_request(connection, f"SELECT * FROM {table};")
        out_file_location = input('Enter output file location:')
        if os.path.isfile(out_file_location) is True:
            ans = input(f"{out_file_location} file exists. Do you want to replace it? (y/n)")
            if ans == 'n':
                sys.exit(0)
        out_workbook = xlsx.Workbook(out_file_location)
        worksheet = out_workbook.add_worksheet()
        for i, row in enumerate(col_names):
            for item in row:
                worksheet.write(0, i, item)
        for i, row in enumerate(data):
            for j, item in enumerate(row):
                worksheet.write(i+1, j, item)
        out_workbook.close()
        print('Done')
elif operation == 'Excel -> sqlite':
    print("""Be careful. 
The upper left corner of your table must be on A1 cell. 
Otherwise, script will be finished with an error.""")
    in_file_location = input('Enter Excel file location:')
    if os.path.isfile(in_file_location) is False:
        print('There is no such file.')
        sys.exit(1)
    else:
        try:
            in_workbook = pyxl.load_workbook(in_file_location)
        except Exception:
            print("This is not Excel file.")
            sys.exit(1)
        if len(in_workbook.sheetnames) > 1:
            sheet_name = input("""Your file has several worksheets. 
Enter name of sheet you want to parse:""")
            try:
                worksheet = in_workbook[sheet_name]
            except Exception:
                print(f"There is no {sheet_name}.")
                sys.exit(1)
        else:
            worksheet = in_workbook.active
        values = []
        for i in range(1, worksheet.max_row + 1):
            row = []
            for j in range(1, worksheet.max_column + 1):
                row.append(worksheet.cell(row=i, column=j).value)
            values.append(row)
        columns = values[0]
        values.pop(0)

        out_file_location = input('Enter output file location:')
        if os.path.isfile(out_file_location) is True:
            ans2 = input(f"{out_file_location} file exists. Data will be written in a new table (y/n).")
            if ans2 == 'n':
                sys.exit(0)
        connection = connect_to_db(out_file_location)
        table_name = input('Enter name of the table to save:')
        change_request(connection, create_query(table_name, columns))
        for v in values:
            change_request(connection, insert_query(table_name, columns, v))
        print('Done')
else:
    print("You didn't specify the operation.")
    sys.exit(1)
